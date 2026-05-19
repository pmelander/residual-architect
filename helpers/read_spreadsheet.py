#!/usr/bin/env python3
"""
Excel/CSV Reader Utility for Residual Architecture Skill Set

Reads Excel (.xlsx, .xls) and CSV files and outputs markdown tables.
Used by the excel-reader skill in skills/utilities/

Usage:
    python read_spreadsheet.py <file> [--sheet <name>] [--rows <max>] [--list-sheets]

Examples:
    python read_spreadsheet.py data.xlsx
    python read_spreadsheet.py data.xlsx --sheet "Sheet2"
    python read_spreadsheet.py data.xlsx --rows 10
    python read_spreadsheet.py data.xlsx --list-sheets
    python read_spreadsheet.py data.csv
"""

import argparse
import csv
import os
import sys
from pathlib import Path


def check_dependencies():
    """Check if required dependencies are available."""
    missing = []

    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")

    if missing:
        print("ERROR: Missing required dependencies:", file=sys.stderr)
        for dep in missing:
            print(f"  - {dep}", file=sys.stderr)
        print("\nInstall with: pip install " + " ".join(missing), file=sys.stderr)
        sys.exit(1)


def get_file_size_mb(file_path):
    """Get file size in megabytes."""
    return os.path.getsize(file_path) / (1024 * 1024)


def sanitize_cell_value(value):
    """Sanitize cell value for markdown output."""
    if value is None:
        return ""

    # Convert to string
    text = str(value).strip()

    # Escape markdown special characters in cells
    text = text.replace("|", "\\|")
    text = text.replace("`", "\\`")

    # Remove control characters
    text = "".join(char for char in text if ord(char) >= 32 or char in ['\n', '\t'])

    # Replace newlines/tabs with spaces for single-line cells
    text = text.replace('\n', ' ').replace('\t', ' ')

    # Limit length to avoid huge cells
    if len(text) > 200:
        text = text[:197] + "..."

    return text


def read_excel_file(file_path, sheet_name=None, max_rows=None):
    """
    Read Excel file and return data as list of lists.

    Args:
        file_path: Path to Excel file
        sheet_name: Specific sheet to read (default: active sheet)
        max_rows: Maximum rows to read (default: all)

    Returns:
        tuple: (data, headers, sheet_name)
    """
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found.", file=sys.stderr)
            print(f"Available sheets: {', '.join(workbook.sheetnames)}", file=sys.stderr)
            sys.exit(1)
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
        sheet_name = sheet.title

    # Read data
    data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if max_rows and i >= max_rows:
            break
        data.append([sanitize_cell_value(cell) for cell in row])

    if not data:
        return [], [], sheet_name

    # First row is headers
    headers = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []

    return rows, headers, sheet_name


def read_csv_file(file_path, max_rows=None):
    """
    Read CSV file and return data as list of lists.

    Args:
        file_path: Path to CSV file
        max_rows: Maximum rows to read (default: all)

    Returns:
        tuple: (data, headers, filename)
    """
    data = []

    with open(file_path, 'r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if max_rows and i >= max_rows:
                break
            data.append([sanitize_cell_value(cell) for cell in row])

    if not data:
        return [], [], Path(file_path).stem

    headers = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []

    return rows, headers, Path(file_path).stem


def format_as_markdown_table(rows, headers):
    """
    Format data as markdown table.

    Args:
        rows: List of lists (data rows)
        headers: List of header values

    Returns:
        str: Markdown table
    """
    if not headers:
        return "_(Empty spreadsheet)_"

    # Calculate column widths
    col_widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(str(cell)))

    # Build table
    lines = []

    # Header row
    header_line = "| " + " | ".join(
        str(h).ljust(col_widths[i]) for i, h in enumerate(headers)
    ) + " |"
    lines.append(header_line)

    # Separator row
    sep_line = "| " + " | ".join(
        "-" * col_widths[i] for i in range(len(headers))
    ) + " |"
    lines.append(sep_line)

    # Data rows
    for row in rows:
        # Pad row if it has fewer columns than headers
        padded_row = row + [""] * (len(headers) - len(row))
        row_line = "| " + " | ".join(
            str(cell).ljust(col_widths[i]) for i, cell in enumerate(padded_row[:len(headers)])
        ) + " |"
        lines.append(row_line)

    return "\n".join(lines)


def list_sheets(file_path):
    """List all sheets in an Excel file."""
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets in {Path(file_path).name}:")
    for i, sheet_name in enumerate(workbook.sheetnames, 1):
        sheet = workbook[sheet_name]
        row_count = sheet.max_row
        col_count = sheet.max_column
        print(f"  {i}. {sheet_name} ({row_count} rows × {col_count} columns)")


def main():
    parser = argparse.ArgumentParser(
        description="Read Excel/CSV files and output markdown tables"
    )
    parser.add_argument("file", help="Path to Excel (.xlsx, .xls) or CSV file")
    parser.add_argument("--sheet", help="Sheet name to read (Excel only)")
    parser.add_argument("--rows", type=int, help="Maximum rows to read")
    parser.add_argument("--list-sheets", action="store_true",
                        help="List available sheets and exit")

    args = parser.parse_args()

    # Validate file exists
    if not os.path.exists(args.file):
        print(f"ERROR: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    # Check file size
    file_size_mb = get_file_size_mb(args.file)
    if file_size_mb > 10:
        print(f"ERROR: File too large ({file_size_mb:.1f} MB). Maximum 10 MB.", file=sys.stderr)
        sys.exit(1)
    elif file_size_mb > 5:
        print(f"WARNING: Large file ({file_size_mb:.1f} MB). This may take a while...", file=sys.stderr)

    # Determine file type
    file_ext = Path(args.file).suffix.lower()

    if file_ext in ['.xlsx', '.xls']:
        check_dependencies()

        if args.list_sheets:
            list_sheets(args.file)
            return

        rows, headers, sheet_name = read_excel_file(args.file, args.sheet, args.rows)

        print(f"# {Path(args.file).name}")
        print(f"**Sheet:** {sheet_name}")
        print(f"**Rows:** {len(rows)}")
        print()
        print(format_as_markdown_table(rows, headers))

        if args.rows and len(rows) >= args.rows:
            print()
            print(f"_(Showing first {args.rows} rows)_")

    elif file_ext == '.csv':
        if args.list_sheets:
            print("CSV files don't have sheets.", file=sys.stderr)
            sys.exit(1)

        if args.sheet:
            print("WARNING: --sheet ignored for CSV files", file=sys.stderr)

        rows, headers, filename = read_csv_file(args.file, args.rows)

        print(f"# {Path(args.file).name}")
        print(f"**Rows:** {len(rows)}")
        print()
        print(format_as_markdown_table(rows, headers))

        if args.rows and len(rows) >= args.rows:
            print()
            print(f"_(Showing first {args.rows} rows)_")

    else:
        print(f"ERROR: Unsupported file type: {file_ext}", file=sys.stderr)
        print("Supported: .xlsx, .xls, .csv", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
